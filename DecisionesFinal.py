from openpyxl import load_workbook
def Decisiones_Fin(Decision,Correcta,Nombre,Respuestas,Premio):
    if (Decision == 's'):
        print('Fin')
        wb3 = load_workbook('players.xlsx')
        hoja2 = wb3.active
        ws3 = wb3['Hoja1']
        maxs3 = ws3.max_row
        hoja2['A%d' % (maxs3 + 1)] = Nombre
        hoja2['B%d' % (maxs3 + 1)] = Premio
        wb3.save('player.xlsx')
        return (Premio, Decision)
    else:
        print('Cual es tu respuesta')
        Respondio = input();
        i = 0;
        while (i < len(Respuestas)):
            RespuestaS = Respuestas[i];
            if (RespuestaS == Respondio):
                NumeroRespuesta = i;
            i += 1;
        if (Correcta[NumeroRespuesta] == '1'):
            Premio=Premio+3000
            print('Ganaste 3000 pesos!')
            print('Ganaste!!! :D')
            print('El premio acumulado es: 6800 pesos!')
            print('')
            print('Gracias por jugar Preguntas y Respuestas!')
            print('')

            wb3 = load_workbook('players.xlsx')
            hoja2 = wb3.active
            ws3 = wb3['Hoja1']
            maxs3 = ws3.max_row
            hoja2['A%d' % (maxs3 + 1)] = Nombre
            hoja2['B%d' % (maxs3 + 1)] = Premio
            wb3.save('players.xlsx')
            Premio = 0
            print('Has ganado :D ')
            return (Premio, Decision)
        else:
            wb3 = load_workbook('players.xlsx')
            hoja2 = wb3.active
            ws3 = wb3['Hoja1']
            Premio = 0
            maxs3 = ws3.max_row
            hoja2['A%d' % (maxs3 + 1)] = Nombre
            hoja2['B%d' % (maxs3 + 1)] = Premio
            wb3.save('players.xlsx')
    return (Premio, Decision)